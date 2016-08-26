from __future__ import unicode_literals
import frappe
import shutil
import json
import os
from frappe.utils import get_datetime, now
import subprocess
from openpyxl import load_workbook


def get_central_delivery():
	central_delivery = frappe.db.sql(""" select distinct usr.email from `tabUser` usr 
								left join `tabUserRole` usr_role 
								on usr_role.parent = usr.name
								where usr.name != "Administrator"
								and usr_role.role = "Central Delivery"  """, as_dict=1)
	central_delivery = [user.get("email") for user in central_delivery if user.get("email")]
	return central_delivery



def make_login_log():	
	log = frappe.new_doc("Login Log")
	log.user = frappe.session.user
	log.login_time = now()
	log.sid = frappe.session.sid
	log.save(ignore_permissions=True)


def update_login_log():
	log = frappe.db.get_value("Login Log", {"sid":frappe.session.sid}, "name")
	if log:
		frappe.db.sql("""  update `tabLogin Log` set log_out_time = %s where name = %s """, (now(), log) )
		frappe.db.commit()


def get_mycfo_users():
	mycfo_user_list = []
	cust_user = frappe.db.sql(""" select distinct usr.name from `tabUser` usr 
								left join `tabUserRole` usr_role 
								on usr_role.parent = usr.name
								where usr.name != "Administrator" 
								and usr_role.role = "Customer"
								and usr.enabled = 1 """, as_dict=1)
	cust_user = [user.get("name") for user in cust_user if user.get("name")]
	cust_user = ','.join('"{0}"'.format(w) for w in cust_user)
	
	mycfo_users = frappe.db.sql(""" select distinct usr.name from `tabUser` usr 
								left join `tabUserRole` usr_role 
								on usr_role.parent = usr.name
								where usr.name != "Administrator"
								and usr.name not in (%s)
								and usr_role.role = "Mycfo User"
								and usr.enabled = 1 """%(cust_user), as_dict=1)
	mycfo_users = [user.get("name") for user in mycfo_users if user.get("name")]
			
	return mycfo_users


def init_ip_file_conversion():
	conv_details = frappe.get_all("IP File Converter", fields=["*"], filters={"converter_status":0, "attempt_count":["<=", "7"]})
	status_list = []
	errors = 0

	for row in conv_details:
		my_dict = {}
		try:
			if row.get("file_extension") == "xlsm":
				convert_xlsm_to_xlsx(row.get("xlsm_path"), row.get("file_path"))

			if os.path.isdir(row.get("dir_path", "")):
				shutil.rmtree(row.get("dir_path", ""), ignore_errors=True)

			args = json.loads(row.get("command"))	
			subprocess.check_call(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
			ip_file_cond = " file_viewer_path = '%s' "%row.get("file_viewer_path")
			update_ip_file(row.get("ip_file"), ip_file_cond)
			my_dict = {"converter_status":1, "attempt_count":int(row.get("attempt_count",0)) + 1}
			status_list.append({"ip_file":row.get("ip_file"), "status":"Conversion Success", "errors":""})
			if os.path.isfile(row.get("edited_file_path")):
				os.remove(row.get("edited_file_path"))
		except Exception, e:
			my_dict = {"traceback":frappe.get_traceback(), "attempt_count":int(row.get("attempt_count",0)) + 1}
			status_list.append({"ip_file":row.get("ip_file"), "status":"Conversion Unsuccessful", "errors":e.message})
			errors += 1
		finally:
			update_ip_file_converter(row.get("name"), my_dict)

	create_ip_file_converter_log(len(conv_details), errors, status_list)		

				

def update_ip_file(ip_file, ip_file_cond):
	query = """ update `tabIP File` set  {0} where name = '{1}' """.format(ip_file_cond, ip_file)
	frappe.db.sql(query)
	frappe.db.commit()

def update_ip_file_converter(ipc_name, ipc_dict):
	ipc = frappe.get_doc("IP File Converter", ipc_name)
	ipc.update(ipc_dict)
	ipc.save(ignore_permissions=True)

def create_ip_file_converter_log(total_count, error_count, converter_log_dict):
	log = frappe.new_doc("IP File Converter Log")
	log.total_count = total_count
	log.error_count = error_count
	log.converter_log_dict = json.dumps(converter_log_dict)
	log.save(ignore_permissions=True)


def convert_xlsm_to_xlsx(xlsm_path, xls_path):
	wb = load_workbook(xlsm_path)
	ws = wb.active
	ws['D2'] = 42
	wb.save(xls_path)